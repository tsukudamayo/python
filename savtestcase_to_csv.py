#!/usr/bin/python3
# _*_ coding: utf-8 _*_

from testcase import sav_testcase
from openpyxl import load_workbook
import sys

wb = load_workbook(str(sys.argv[1]))
#m = sys.argv[2]
#n = sys.argv[3]

all_sheetnames = wb.get_sheet_names()
print(all_sheetnames)
sheetnames = all_sheetnames[int(10):int(16)]
print(sheetnames)

for sheetname in sheetnames:
    sav_testcase(str(sys.argv[1]), sheetname, str(sheetname) + '.csv')
