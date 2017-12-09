#!/usr/bin/python3
# _*_ coding: utf-8 _*_

from testcase import sim_testcase
from openpyxl import load_workbook
import sys

wb = load_workbook(str(sys.argv[1]))
n = sys.argv[2]

all_sheetnames = wb.get_sheet_names()
print(all_sheetnames)
sheetnames = all_sheetnames[int(n):]
print(sheetnames)

for sheetname in sheetnames:
    sim_testcase(str(sys.argv[1]), sheetname, str(sheetname) + '.csv')
